#!/usr/bin/env python3
"""
build_excel.py — Regenerate Europe-2026-VRFamilyTravels-Final.xlsx from europe-2026.json
Run after ANY data change: python3 build_excel.py
"""
import json, re, openpyxl, shutil, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('data/europe-2026.json') as f:
    d = json.load(f)

def cl(t): return re.sub(r'<[^>]+>', '', str(t or '')).strip()
def se(t): return re.sub(r'[\U00010000-\U0010ffff]', '', str(t or '')).strip()

INK='1A1A2E'; GOLD='C9A84C'; WHITE='FFFFFF'; CREAM='FAF8F3'
SLATE='6B7280'; SAGE='4A7C2F'; RUST='C0392B'; AMBER='92400E'
DAY_BG='2D3561'; TL_BG='1D2645'; KD_BG='FFFBEB'; TIP_BG='F8F8FF'
TYPE_BG={'DRIVE':'EFF6FF','HIKE':'F0F7EC','MEAL':'FEF9EE','HOTEL':'F5F0FF',
         'WARN':'FEF2F2','STOP':'FFFFFF'}
TYPE_FG={'DRIVE':'1D4ED8','HIKE':SAGE,'MEAL':AMBER,'HOTEL':'5B21B6','WARN':RUST,'STOP':'111111'}
TYPE_LBL={'DRIVE':'Drive','HIKE':'Hike','MEAL':'Eat','HOTEL':'Stay','WARN':'Alert','STOP':'Stop'}
STATUS_BG={'done':'F0F7EC','urgent':'FEF2F2','pending':'FFFBEB'}
STATUS_FG={'done':SAGE,'urgent':RUST,'pending':AMBER}
STATUS_LBL={'done':'Done','urgent':'Urgent','pending':'Pending'}

def thin(c='E0E0E0'):
    s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)

def sc(ws,r,c,v='',bg=None,fg='111111',bold=False,sz=10,wrap=True,align='left',italic=False,valign='top'):
    cell=ws.cell(row=r,column=c,value=str(v) if v is not None else '')
    cell.font=Font(name='Arial',color=fg,bold=bold,size=sz,italic=italic)
    if bg: cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal=align,vertical=valign,wrap_text=wrap)
    cell.border=thin()
    return cell

def rh(ws,r,h): ws.row_dimensions[r].height=h

def classify(text):
    t=text.lower()
    if re.search(r'drive |fly |flight|depart|checkout|check-out|pick up',t): return 'DRIVE'
    if re.search(r'\bhike\b|trail|loop|climb|summit|descent|ascent',t): return 'HIKE'
    if re.search(r'lunch|dinner|breakfast|cafe|rifugio|restaurant|grocery|supermarket',t): return 'MEAL'
    if re.search(r'check.in|night [0-9]|lodge|apartment|camping|gasthof|hotel\b',t): return 'HOTEL'
    if re.search(r'alert|closed|critical|2026 alert',t): return 'WARN'
    return 'STOP'

SKIP_ITEMS=[r'night [0-9].{0,10}already checked in',r'^austrian vignette still valid',
            r'^spend remaining swiss francs',r'^gondola down.*drive back',
            r'^2:30pm.*back at landhaus',r'^pack bags']
def skip_item(text):
    return any(re.search(p,text.lower().strip()) for p in SKIP_ITEMS)

# Map links
map_links={}
for day in d['days']:
    links=day.get('mapLinks',[])
    if not links: continue
    if day['dayLabel']=='1–2':
        for ml in links:
            lbl=ml.get('label','').lower()
            if 'day 1' in lbl: map_links['Jun 20']=ml.get('url','')
            elif 'day 2' in lbl: map_links['Jun 21']=ml.get('url','')
    else:
        primary=next((ml for ml in links if 'hofer' not in ml.get('label','').lower()),links[0])
        map_links[day['date']]=primary.get('url','')

wb=Workbook()

# ═══ TAB 1: DETAILED ITINERARY ════════════════════════════════════════════
ws1=wb.active; ws1.title='Detailed Itinerary'
ws1.sheet_view.showGridLines=False; ws1.freeze_panes='A3'
for col,w in zip('ABCDEF',[7,10,8,5,72,22]): ws1.column_dimensions[col].width=w
ws1.merge_cells('A1:F1')
c1=ws1.cell(row=1,column=1,value='Europe 2026 - Detailed Itinerary  |  VR Family Travels  |  Jun 19 - Jul 5')
c1.font=Font(name='Arial',bold=True,color=WHITE,size=12)
c1.fill=PatternFill('solid',start_color=INK)
c1.alignment=Alignment(horizontal='left',vertical='center'); rh(ws1,1,26)
for ci,v in enumerate(['Day','Date','Type','Opt','Activity / Details','Key Info'],1):
    c2=ws1.cell(row=2,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=INK,size=9)
    c2.fill=PatternFill('solid',start_color=GOLD)
    c2.alignment=Alignment(horizontal='center',vertical='center')
    c2.border=thin(); rh(ws1,2,16)

def add_day(ws,row,label,date,title,items,timeline,kd,tags,drive,dtime):
    ws.merge_cells(f'A{row}:F{row}')
    hdr=f'DAY {label}  |  {date}  |  {title}'
    if drive: hdr+=f'  |  {drive}  {dtime}'
    c=ws.cell(row=row,column=1,value=hdr)
    c.font=Font(name='Arial',bold=True,color=WHITE,size=10)
    c.fill=PatternFill('solid',start_color=DAY_BG)
    c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=thin('3D4C7A'); rh(ws,row,20); row+=1
    if timeline:
        ws.merge_cells(f'A{row}:F{row}')
        tc=ws.cell(row=row,column=1,value=f'Timeline: {se(timeline)}')
        tc.font=Font(name='Arial',color=GOLD,size=8.5,italic=True)
        tc.fill=PatternFill('solid',start_color=TL_BG)
        tc.alignment=Alignment(horizontal='left',vertical='center')
        tc.border=thin(); rh(ws,row,15); row+=1
    if kd:
        kds=se('  |  '.join(kd) if isinstance(kd,list) else kd)
        ws.merge_cells(f'A{row}:F{row}')
        kc=ws.cell(row=row,column=1,value=f'Key: {kds}')
        kc.font=Font(name='Arial',color=AMBER,size=8.5)
        kc.fill=PatternFill('solid',start_color=KD_BG)
        kc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
        kc.border=thin(); rh(ws,row,22); row+=1
    for item in (items or []):
        text=cl(item)
        if not text or skip_item(text): continue
        typ=classify(text)
        is_opt=text.lower().startswith('optional') or ('+' in item[:5] and 'optional' in text.lower())
        bg=TYPE_BG.get(typ,'FFFFFF'); fg=TYPE_FG.get(typ,'111111')
        if is_opt: bg='F9F9F9'; fg='888888'
        keys=' / '.join((re.findall(r'EUR\s?[\d.,]+|CHF\s?[\d.]+|\$[\d.]+|€[\d.,]+|~?\d+\s?min|~?\d+\s?hr|~?\d+\s?km',text,re.I) or [])[:3])[:28]
        sc(ws,row,1,'',bg=bg,sz=8,align='center')
        sc(ws,row,2,date,bg=bg,fg=SLATE,sz=8,align='center')
        sc(ws,row,3,TYPE_LBL.get(typ,typ),bg=bg,fg=fg,sz=8,align='center',bold=(typ in ('WARN','HOTEL','DRIVE')))
        sc(ws,row,4,'opt' if is_opt else '',bg=bg,fg=SLATE,sz=8,align='center')
        sc(ws,row,5,se(text)[:220],bg=bg,fg=RUST if typ=='WARN' else ('666666' if is_opt else '111111'),sz=10)
        sc(ws,row,6,keys,bg=bg,fg=SLATE,sz=9)
        rh(ws,row,max(18,min(50,12+len(text)//5))); row+=1
    for tag in (tags or []):
        txt=se(cl(tag.get('text','')))
        if len(txt)<5: continue
        ws.merge_cells(f'C{row}:F{row}')
        for ci2 in [1,2]: sc(ws,row,ci2,'',bg=TIP_BG)
        ws.cell(row=row,column=2).value='Tip'
        ws.cell(row=row,column=2).font=Font(name='Arial',color='4338CA',size=8,bold=True)
        ws.cell(row=row,column=2).fill=PatternFill('solid',start_color=TIP_BG)
        ws.cell(row=row,column=2).alignment=Alignment(horizontal='center',vertical='top')
        ws.cell(row=row,column=2).border=thin()
        tip=ws.cell(row=row,column=3,value=txt[:240])
        tip.font=Font(name='Arial',color=SLATE,size=8.5,italic=True)
        tip.fill=PatternFill('solid',start_color=TIP_BG)
        tip.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
        tip.border=thin(); rh(ws,row,max(14,min(34,10+len(txt)//7))); row+=1
    for ci2 in range(1,7):
        ws.cell(row=row,column=ci2).fill=PatternFill('solid',start_color='E8E4DC')
        ws.cell(row=row,column=ci2).border=thin()
    rh(ws,row,3); row+=1
    return row

row=3
for day in d['days']:
    tl=day.get('timeline',''); kd=day.get('keyDecisions',[])
    dr=day['summary'].get('drive',''); dt=day['summary'].get('driveTime','')
    if day['dayLabel']=='1–2':
        row=add_day(ws1,row,'1','Jun 20','Iceland Day 1 - South Coast Gateway',
            day['sections'][0]['items'],
            'Day 1: 10:30am KEF > Selfoss > Geysir > Seljalandsfoss+Gljufrabui > Skogafoss > Reynisfjara > 7:30pm Vik',
            ['Selfoss lunch+grocery first','Gljufrabui MUST - waterproofs','Reynisfjara 6pm golden light'],
            day.get('tags',[]),'~180 km','~2.5h driving')
        row=add_day(ws1,row,'2','Jun 21','Iceland Day 2 - Glacier Walk + Jokulsarlon',
            day['sections'][1]['items'],
            'Day 2: 7am > Dyrholaey > Glacier walk > Skaftafell > Jokulsarlon > Diamond Beach > 2pm drive > 4pm Canyon > Vik > 8pm BergOne',
            ['Glacier walk free - kids touch ice','2pm departure firm','Canyon on return at 4pm'],
            day['sections'][1].get('tags',[]),'~480 km','~5h driving')
    else:
        all_items=sum([s['items'] for s in day['sections']],[])
        all_tags=day.get('tags',[])+sum([s.get('tags',[]) for s in day['sections']],[])
        row=add_day(ws1,row,day['dayLabel'],day['date'],se(day['title']),all_items,
            se(tl) if tl else '',kd,all_tags,dr,dt)

# ═══ TAB 2: TRIP OVERVIEW & TRACKER ══════════════════════════════════════
ws2=wb.create_sheet('Trip Overview & Tracker')
ws2.sheet_view.showGridLines=False; ws2.freeze_panes='A3'
for col,w in zip('ABCDEFGH',[8,10,26,28,56,18,12,38]): ws2.column_dimensions[col].width=w
ws2.merge_cells('A1:H1')
t=ws2.cell(row=1,column=1,value='Europe 2026 - Trip Overview & Live Tracker  |  Jun 19 - Jul 5')
t.font=Font(name='Arial',bold=True,color=WHITE,size=12)
t.fill=PatternFill('solid',start_color=INK)
t.alignment=Alignment(horizontal='left',vertical='center'); rh(ws2,1,26)
for ci,v in enumerate(['Day','Date','From','Night Stay','Highlights','Drive','Map','Notes'],1):
    c2=ws2.cell(row=2,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=INK,size=9)
    c2.fill=PatternFill('solid',start_color=GOLD)
    c2.alignment=Alignment(horizontal='center',vertical='center')
    c2.border=thin(); rh(ws2,2,16)

COUNTRY_BG={'Jun 20':'EFF9FF','Jun 21':'EFF9FF','Jun 22':'F0FFF0','Jun 23':'F0FFF0',
            'Jun 24':'FFFFF0','Jun 25':'FFFFF0','Jun 26':'FFF5EE','Jun 27':'FFF5EE',
            'Jun 28':'FFF5EE','Jun 29':'FFF0F5','Jun 30':'FFF0F5','Jul 1':'FFF5EE',
            'Jul 2':'FFF5EE','Jul 3':'FFF5EE','Jul 4':'F0FFF0','Jul 5':'EFF9FF'}

HIGHLIGHTS={
    'Jun 20':'KEF pickup · Selfoss lunch+grocery · Geysir/Strokkur · Seljalandsfoss+Gljufrabui · Skogafoss · Reynisfjara golden hour',
    'Jun 21':'Dyrholaey puffins · Glacier walk (free) · Jokulsarlon shore · Diamond Beach · Fjadrargljufur Canyon 4pm · Vik dinner',
    'Jun 22':'Fly KEF>ZRH · Europcar pickup 1:30pm · Spiez Castle · Coop Lauterbrunnen · Staubbach Falls evening',
    'Jun 23':'Oeschinensee gondola · Heuberg Ridge Trail #8 · Grindelwald First · Cliff Walk · Bachalpsee',
    'Jun 24':'Trummelbach Falls · Vaduz Liechtenstein (Country #3) · A8 construction warning · Berchtesgaden',
    'Jun 25':'Konigssee boat to Salet · Obersee+Rothbachfall · St. Bartholoma lunch · Eagles Nest (clear) or Salt Mine (cloudy) · Rossfeld · Zauberwald',
    'Jun 26':'Liechtensteinklamm 9am · Salzburg P+R South · Hohensalzburg · Wolfgangsee ferry · Hallstatt golden hour',
    'Jun 27':'Krippenstein 8:45am · 5 Fingers viewpoint · Gosausee Mirror Lake · Gosaukammbahn · Hallstatt 4:30pm',
    'Jun 28':'Grossglockner 46.50 EUR · Edelweiss-Spitze 2571m · KFJ-Hohe marmots · Heiligenblut church photo · Sillian',
    'Jun 29':'Braies 8:45am · Misurina mirror coffee · Cadini Trail 117 · Rifugio Auronzo lunch · Tre Cime walk · Gardena Pass · Collalbo 5pm',
    'Jun 30':'Seceda 8:45am · Razor Edge Trail 1 · Santa Maddalena (2026 barrier) · St. Johann Ranui · Adolf Munkel',
    'Jul 1': '9am St. Valentin gate CRITICAL · Bullaccia loop · Hexenbanke viewpoint · Tschotsch Alm lunch · Brenner construction',
    'Jul 2': 'Schlegeis toll 17 EUR · Olpererhütte hike · Kebema bridge 10:30am · Swarovski if rain · Heiterwanger See',
    'Jul 3': 'Ehrwalder Almbahn · Seebensee mirror (before 10:30am) · Drachensee climb · Coburger Hutte Kaiserschmarrn',
    'Jul 4': 'Marienbrücke 8:30am FIRST · Neuschwanstein 9:30am · Ebenalp · Aescher before 5pm · return car tonight',
    'Jul 5': 'Hilton shuttle 5:30am (NOT walk) · Terminal 2 Row 3 · FI571 7:55am ZRH>KEF>MCO · home MCO 8:50pm',
}
NIGHT_STAY={
    'Jun 20':'Farmhouse Lodge, Vik, Iceland','Jun 21':'BergOne, Ytri-Njardvik, Iceland',
    'Jun 22':'Camping Jungfrau, Lauterbrunnen, Switzerland','Jun 23':'Camping Jungfrau, Lauterbrunnen, Switzerland',
    'Jun 24':'Apartment, Berchtesgaden, Germany','Jun 25':'Apartment, Berchtesgaden, Germany',
    'Jun 26':'Kaiserblick Goisern, Bad Goisern, Austria','Jun 27':'Kaiserblick Goisern, Bad Goisern, Austria',
    'Jun 28':'Hotel Schwarzer Adler, Sillian, Austria','Jun 29':'Kaiserau 1907, Collalbo, Italy',
    'Jun 30':'Kaiserau 1907, Collalbo, Italy','Jul 1':'Keyone Rooms, Finkenberg, Austria',
    'Jul 2':'Landhaus Bichlbach, Austria','Jul 3':'Landhaus Bichlbach, Austria',
    'Jul 4':'Airport Hotel Zurich, Switzerland','Jul 5':'Home - Orlando, Florida',
}
FROM_MAP={
    'Jun 20':'Keflavik KEF, Iceland','Jun 21':'Farmhouse Lodge, Vik',
    'Jun 22':'Camping Jungfrau / fly KEF>ZRH','Jun 23':'Camping Jungfrau, Lauterbrunnen',
    'Jun 24':'Camping Jungfrau, Lauterbrunnen','Jun 25':'Apartment, Berchtesgaden',
    'Jun 26':'Apartment, Berchtesgaden','Jun 27':'Kaiserblick, Bad Goisern',
    'Jun 28':'Kaiserblick, Bad Goisern','Jun 29':'Hotel Schwarzer Adler, Sillian',
    'Jun 30':'Kaiserau 1907, Collalbo','Jul 1':'Kaiserau 1907, Collalbo',
    'Jul 2':'Keyone Rooms, Finkenberg','Jul 3':'Landhaus Bichlbach',
    'Jul 4':'Landhaus Bichlbach','Jul 5':'Airport Hotel Zurich',
}
DRIVE_MAP={}
for g in d['glance']: DRIVE_MAP[g['date']]=(g.get('drive','')+' '+g.get('driveTime','')).strip()
DRIVE_MAP['Jun 20']='~180 km  ~2.5h'; DRIVE_MAP['Jun 21']='~480 km  ~5h'

all_dates=['Jun 20','Jun 21','Jun 22','Jun 23','Jun 24','Jun 25','Jun 26','Jun 27',
           'Jun 28','Jun 29','Jun 30','Jul 1','Jul 2','Jul 3','Jul 4','Jul 5']
day_nums=['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16']

for i,(date,label) in enumerate(zip(all_dates,day_nums)):
    rn=i+3; bg=COUNTRY_BG.get(date,CREAM)
    sc(ws2,rn,1,f'Day {label}',bg=bg,bold=True,align='center',fg=INK)
    sc(ws2,rn,2,date,bg=bg,bold=True)
    sc(ws2,rn,3,FROM_MAP.get(date,''),bg=bg,fg=SLATE,sz=9)
    sc(ws2,rn,4,NIGHT_STAY.get(date,''),bg=bg)
    sc(ws2,rn,5,HIGHLIGHTS.get(date,''),bg=bg)
    sc(ws2,rn,6,DRIVE_MAP.get(date,''),bg=bg,fg=SLATE,sz=9)
    url=map_links.get(date,'')
    mc=ws2.cell(row=rn,column=7)
    if url:
        mc.value='Open Map'; mc.hyperlink=url
        mc.font=Font(name='Arial',color='1D4ED8',size=9,underline='single',bold=True)
    else:
        mc.value=''; mc.font=Font(name='Arial',color=SLATE,size=9)
    mc.fill=PatternFill('solid',start_color=bg)
    mc.alignment=Alignment(horizontal='center',vertical='top')
    mc.border=thin()
    sc(ws2,rn,8,'',bg=bg); rh(ws2,rn,32)

lr=len(all_dates)+4
ws2.merge_cells(f'A{lr}:H{lr}')
leg=ws2.cell(row=lr,column=1,value='Map opens Google Maps route · Notes: update during trip · Blue=Iceland  Green=Switzerland  Yellow=Germany  Orange=Austria  Pink=Italy')
leg.font=Font(name='Arial',color=SLATE,size=9,italic=True)
leg.alignment=Alignment(horizontal='left',vertical='center'); rh(ws2,lr,16)

# ═══ TAB 3: BOOKING CHECKLIST ═════════════════════════════════════════════
ws3=wb.create_sheet('Booking Checklist')
ws3.sheet_view.showGridLines=False; ws3.freeze_panes='A3'
for col,w in zip('ABCDE',[12,52,38,42,5]): ws3.column_dimensions[col].width=w
ws3.merge_cells('A1:E1')
t3=ws3.cell(row=1,column=1,value='Europe 2026 - Booking Checklist')
t3.font=Font(name='Arial',bold=True,color=WHITE,size=12)
t3.fill=PatternFill('solid',start_color=INK)
t3.alignment=Alignment(horizontal='left',vertical='center'); rh(ws3,1,26)
for ci,v in enumerate(['Status','Item','Reference','Deadline / Notes','Done'],1):
    c2=ws3.cell(row=2,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=INK,size=9)
    c2.fill=PatternFill('solid',start_color=GOLD)
    c2.alignment=Alignment(horizontal='center',vertical='center')
    c2.border=thin(); rh(ws3,2,16)
for i,item in enumerate(d['checklist']):
    st=item.get('status','pending'); bg=STATUS_BG.get(st,WHITE); rn=i+3
    sc(ws3,rn,1,STATUS_LBL.get(st,st),bg=bg,fg=STATUS_FG.get(st,SLATE),bold=True,align='center')
    sc(ws3,rn,2,se(item.get('label','')),bg=bg)
    sc(ws3,rn,3,se(item.get('ref','') or ''),bg=bg,fg=SLATE,sz=9)
    sc(ws3,rn,4,se(item.get('deadline','') or ''),bg=bg,fg=SLATE,sz=9)
    cb=ws3.cell(row=rn,column=5,value='Y' if st=='done' else '')
    cb.font=Font(name='Arial',size=10,color=SAGE if st=='done' else SLATE,bold=(st=='done'))
    cb.fill=PatternFill('solid',start_color=bg)
    cb.alignment=Alignment(horizontal='center',vertical='center')
    cb.border=thin(); rh(ws3,rn,26)

# ═══ TAB 4: ACCOMMODATION ═════════════════════════════════════════════════
ws4=wb.create_sheet('Accommodation')
ws4.sheet_view.showGridLines=False; ws4.freeze_panes='A3'
for col,w in zip('ABCDE',[8,14,32,50,32]): ws4.column_dimensions[col].width=w
ws4.merge_cells('A1:E1')
t4=ws4.cell(row=1,column=1,value='Europe 2026 - Accommodation')
t4.font=Font(name='Arial',bold=True,color=WHITE,size=12)
t4.fill=PatternFill('solid',start_color=INK)
t4.alignment=Alignment(horizontal='left',vertical='center'); rh(ws4,1,26)
for ci,v in enumerate(['Nights','Dates','Property','Address','Phone / Contact'],1):
    c2=ws4.cell(row=2,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=INK,size=9)
    c2.fill=PatternFill('solid',start_color=GOLD)
    c2.alignment=Alignment(horizontal='center',vertical='center')
    c2.border=thin(); rh(ws4,2,16)
for i,a in enumerate([x for x in d['accommodation'] if x.get('status')!='cancelled']):
    bg='F0F7EC' if i%2==0 else WHITE; rn=i+3
    sc(ws4,rn,1,a.get('nights',''),bg=bg,align='center',bold=True)
    sc(ws4,rn,2,a.get('dates',''),bg=bg,bold=True)
    sc(ws4,rn,3,se(a.get('property','')),bg=bg,bold=True)
    sc(ws4,rn,4,a.get('address',''),bg=bg,fg=SLATE,sz=9)
    sc(ws4,rn,5,a.get('phone',''),bg=bg,fg=SLATE,sz=9)
    rh(ws4,rn,26)

# ═══ TAB 5: BUDGET ════════════════════════════════════════════════════════
ws5=wb.create_sheet('Budget')
ws5.sheet_view.showGridLines=False
for col,w in zip('ABCD',[32,14,14,44]): ws5.column_dimensions[col].width=w
ws5.merge_cells('A1:D1')
t5=ws5.cell(row=1,column=1,value='Europe 2026 - Budget Tracker')
t5.font=Font(name='Arial',bold=True,color=WHITE,size=12)
t5.fill=PatternFill('solid',start_color=INK)
t5.alignment=Alignment(horizontal='left',vertical='center'); rh(ws5,1,26)
for ci,v in enumerate(['Category','Estimated','Actual Spent','Notes'],1):
    c2=ws5.cell(row=2,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=INK,size=9)
    c2.fill=PatternFill('solid',start_color=GOLD)
    c2.alignment=Alignment(horizontal='center',vertical='center')
    c2.border=thin(); rh(ws5,2,16)
budget=[
    ('Flights x4 (confirmed)','$4,088','$4,088','Icelandair A77762'),
    ('Accommodation 15 nights (confirmed)','$3,032','$3,032','All 10 properties'),
    ('Car Rental ZRH + KEF (confirmed)','$1,182','$1,182','Europcar 738796923 + Budget #42334482US1'),
    ('Food & Drinks','$1,850','','Update daily'),
    ('Activities & Entry Fees','$1,100','','Cable cars, castle entries'),
    ('Fuel & Tolls','$1,000','','Grossglockner 46.50 EUR, Schlegeis 17 EUR, Auronzo 40 EUR'),
    ('Shopping & Misc','$400','',''),
]
for i,(cat,est,act,note) in enumerate(budget):
    is_conf='confirmed' in cat; bg='F0F7EC' if is_conf else (CREAM if i%2==0 else WHITE); rn=i+3
    sc(ws5,rn,1,cat,bg=bg,bold=is_conf)
    sc(ws5,rn,2,est,bg=bg,fg=SLATE,align='right')
    sc(ws5,rn,3,act or '',bg=bg,fg=SAGE,bold=bool(act),align='right')
    sc(ws5,rn,4,note,bg=bg,fg=SLATE,sz=9); rh(ws5,rn,22)
rn=len(budget)+3
for ci,(v,fg2,bg2) in enumerate([('TOTAL CONFIRMED',WHITE,INK),('$8,302',GOLD,INK),
    ('',GOLD,INK),('vs est. ~$11,600-$12,900',CREAM,INK)],1):
    c2=ws5.cell(row=rn,column=ci,value=v)
    c2.font=Font(name='Arial',bold=True,color=fg2,size=11)
    c2.fill=PatternFill('solid',start_color=bg2)
    c2.alignment=Alignment(horizontal='right' if ci in [2,3] else 'left',vertical='center')
    c2.border=thin(); rh(ws5,rn,24)

out='Europe-2026-VRFamilyTravels-Final.xlsx'
wb.save(out)
print(f"✓ {row-3} itinerary rows · saved: {out}")
